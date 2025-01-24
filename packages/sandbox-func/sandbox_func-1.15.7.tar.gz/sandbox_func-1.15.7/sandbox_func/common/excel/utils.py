import os
import logging
import shutil

from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from copy import copy
from sandbox_func.common.util.pathutil import get_tmp_path

logger = logging.getLogger(__name__)


def parse_template_data(workbook, temp_params):
    """
    模板变量处理
    :param workbook: 工作表对象
    :param temp_params: 模板参数
    """
    try:
        # 遍历所有工作表
        for sheet in workbook.worksheets:
            rows_to_delete = []  # 存储需要删除的占位符行
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # 普通占位符替换
                        for placeholder, value in temp_params.items():
                            if f"{{{{{placeholder}}}}}" in cell.value:
                                cell.value = cell.value.replace(f"{{{{{placeholder}}}}}", str(value))

                        # 列表占位符替换
                        if cell.value.startswith("{list:") and cell.value.endswith("}"):
                            list_name = cell.value[6:-1]  # 提取 list 名称

                            if list_name in temp_params:
                                start_row = cell.row
                                if len(temp_params[list_name]) > 0:
                                    insert_rows(sheet, start_row, temp_params[list_name])
                                    rows_to_delete.append(start_row)
                                else:
                                    cell.value = ""
                            else:
                                cell.value = ""

            # 删除嵌套占位符行（避免空行残留）
            for row_index in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_index)

        logging.info("模板变量处理完成")
        return True
    except Exception as e:
        print(f"模板变量处理失败: {e}")


def insert_rows(sheet, start_row, data_list):
    num_rows = len(data_list)
    # 记录插入前的合并单元格范围
    original_merged_ranges = list(sheet.merged_cells.ranges)

    # 保存下方原有行样式和合并单元格信息
    original_row_styles = {row: sheet.row_dimensions[row].height for row in range(start_row + 1, sheet.max_row + 1)}

    # 插入新行
    sheet.insert_rows(start_row + 1, amount=num_rows)

    # 获取参考行的样式和信息
    ref_row_height = sheet.row_dimensions[start_row].height
    max_col = sheet.max_column
    ref_cells = [sheet.cell(row=start_row, column=col) for col in range(1, max_col + 1)]

    # 填充新行的数据和样式
    for _index, _item in enumerate(data_list):
        _item.reverse()  # 反转数据
        current_row = start_row + 1 + _index
        sheet.row_dimensions[current_row].height = ref_row_height
        for ref_cell in ref_cells:
            target_cell = sheet.cell(row=current_row, column=ref_cell.column)
            copy_style(ref_cell, target_cell)  # 复制样式
            if isinstance(ref_cell, MergedCell):
                continue
            if _item:
                target_cell.value = _item.pop()

    # 处理合并单元格
    for merged_range in original_merged_ranges:
        if start_row in range(merged_range.min_row, merged_range.max_row + 1):
            for i in range(num_rows):
                offset = i + 1
                new_start_row = merged_range.min_row + offset
                new_end_row = merged_range.max_row + offset
                new_range = f"{get_column_letter(merged_range.min_col)}{new_start_row}:{get_column_letter(merged_range.max_col)}{new_end_row}"
                sheet.merge_cells(new_range)
        if merged_range.min_row > start_row:
            merged_range.shift(0, num_rows - 1)

    for row, height in original_row_styles.items():
        # 计算恢复到新位置的行号
        new_row = row + num_rows - 1
        sheet.row_dimensions[new_row].height = height


def copy_style(ref_cell, target_cell):
    if ref_cell.has_style:
        target_cell.font = copy(ref_cell.font)
        target_cell.border = copy(ref_cell.border)
        target_cell.fill = copy(ref_cell.fill)
        target_cell.number_format = ref_cell.number_format
        target_cell.protection = copy(ref_cell.protection)
        target_cell.alignment = copy(ref_cell.alignment)


def clear_directory(template_file_path):
    try:
        if template_file_path:
            _dir = os.path.dirname(template_file_path)
            logger.debug(f"删除导出目录: {_dir}")
            shutil.rmtree(_dir)
    except Exception as e:
        logger.exception(e)
        logger.debug("删除导出目录失败")


def get_request_tmp_path(request_id: str) -> str:
    return os.path.join(get_tmp_path(), "excel", request_id)
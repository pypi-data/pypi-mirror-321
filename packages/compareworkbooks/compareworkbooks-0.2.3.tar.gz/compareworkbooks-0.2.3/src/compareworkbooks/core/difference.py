import openpyxl

__all__ = ["worksheets", "workbooks", "files"]


def cells(cellA, cellB, *, title, colors=True, values=True):
    lines = list()

    if cellA.coordinate != cellB.coordinate:
        line = "Worksheet %r coordinate: %r vs %r"
        line %= (title, cellA.coordinate, cellB.coordinate)
        lines.append(line)
        return lines

    # Compare cell values
    if values:
        if cellA.value != cellB.value:
            line = "Value in Worksheet %r, Cell %r: %r vs %r"
            line %= (title, cellA.coordinate, cellA.value, cellB.value)
            lines.append(line)

    # Compare fill colors
    # Note: fill.fgColor can be in different formats (e.g., theme-based color),
    # but .rgb is a good quick check for many cases.
    if colors:
        color1 = cellA.fill.fgColor.rgb if cellA.fill.fgColor else None
        color2 = cellB.fill.fgColor.rgb if cellB.fill.fgColor else None
        if color1 != color2:
            line = "Color in Worksheet %r, Cell %r: %r vs %r"
            line %= (title, cellA.coordinate, color1, color2)
            lines.append(line)

    return lines


def worksheets(wsA, wsB, **kwargs):
    # 1) Check sheet names
    if wsA.title != wsB.title:
        return "Worksheets title: %r vs %r" % (wsA.title, wsB.title)

    # 2) Compare cell-by-cell
    max_row = max(wsA.max_row, wsB.max_row)
    max_col = max(wsA.max_column, wsB.max_column)

    lines = []

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cellA = wsA.cell(row=row, column=col)
            cellB = wsB.cell(row=row, column=col)
            lines += cells(cellA, cellB, title=wsA.title, **kwargs)

    return "\n".join(lines).strip()


def workbooks(wbA, wbB, **kwargs):
    """
    Compare two workbooks to check if they are equivalent:
      - same number of sheets
      - each corresponding sheet is equivalent
    """

    # 1) Compare the number of worksheets
    if len(wbA.worksheets) != len(wbB.worksheets):
        return "Number of worksheets: %r vs %r" % (
            len(wbA.worksheets),
            len(wbB.worksheets),
        )

    # 2) Compare sheets one by one
    for wsA, wsB in zip(wbA.worksheets, wbB.worksheets):
        msg = worksheets(wsA, wsB, **kwargs)
        if msg:
            return msg

    return ""


def files(fileA, fileB, **kwargs):
    # Load workbooks
    wbA = openpyxl.load_workbook(fileA)
    wbB = openpyxl.load_workbook(fileB)

    return workbooks(wbA, wbB, **kwargs)

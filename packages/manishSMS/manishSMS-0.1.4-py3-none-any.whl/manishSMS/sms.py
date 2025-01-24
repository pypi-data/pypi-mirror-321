import random
import string
import os
from tabulate import tabulate
from openpyxl import Workbook, load_workbook

database = "school_management.xlsx"


class StudentManagementSystem:
    def __init__(self):
        self.students = []
        self.staffs = []
        self.courses = []
        self.load_from_excel()

    # Loading data from Excel
    def load_from_excel(self):
        if not os.path.exists(database):
            wb = Workbook()
            wb.create_sheet(title="Students")
            wb.create_sheet(title="Staff")
            wb.create_sheet(title="Courses")
            wb.save(database)
            wb.close()
            print("Workbook created with necessary sheets at present working directory.")
            return

        wb = load_workbook(database)
        if "Students" not in wb.sheetnames:
            wb.create_sheet(title="Students")
        if "Staff" not in wb.sheetnames:
            wb.create_sheet(title="Staff")
        if "Courses" not in wb.sheetnames:
            wb.create_sheet(title="Courses")
    
        wb.save(database)  
        wb.close()

        # Reloading the workbook to access sheets
        wb = load_workbook(database)
        self.load_students(wb['Students'])
        self.load_staff(wb['Staff'])
        self.load_courses(wb['Courses'])
        wb.close()

    def load_students(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student = {
                'id': row[0],
                'enrollment': row[1],
                'name': row[2],
                'email': row[3],
                'phone': row[4],
                'address': row[5],
                'course': row[6]
            }
            self.students.append(student)

    def load_staff(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            staff = {
                'id': row[0],
                'name': row[1],
                'email': row[2],
                'phone': row[3],
                'course': row[4]
            }
            self.staffs.append(staff)

    def load_courses(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            course = {
                'id': row[0],
                'name': row[1],
                'assigned_staff': row[2]
            }
            self.courses.append(course)

    # Saving data to Excel
    def save_to_excel(self):
        wb = Workbook()
        self.save_students(wb.create_sheet(title="Students"))
        self.save_staff(wb.create_sheet(title="Staff"))
        self.save_courses(wb.create_sheet(title="Courses"))
        wb.save(database)
        wb.close()

    def save_students(self, sheet):
        headers = ['ID', 'Enrollment', 'Name', 'Email', 'Phone', 'Address', 'Course']
        sheet.append(headers)
        for student in self.students:
            sheet.append([
                student['id'],
                student['enrollment'],
                student['name'],
                student['email'],
                student['phone'],
                student['address'],
                student['course']
            ])

    def save_staff(self, sheet):
        headers = ['ID', 'Name', 'Email', 'Phone', 'Course']
        sheet.append(headers)
        for staff in self.staffs:
            sheet.append([
                staff['id'],
                staff['name'],
                staff['email'],
                staff['phone'],
                staff['course']
            ])

    def save_courses(self, sheet):
        headers = ['ID', 'Name', 'Assigned Staff']
        sheet.append(headers)
        for course in self.courses:
            sheet.append([
                course['id'],
                course['name'],
                course['assigned_staff']
            ])

    # Generating unique ID, enrollment number, Staff ID and Course ID
    def generate_unique_student_id(self):
        while True:
            student_id = random.randint(1000, 9999)
            if not any(student['id'] == student_id for student in self.students):
                return student_id

    def generate_enrollment_number(self):
        while True:
            enrollment = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not any(student['enrollment'] == enrollment for student in self.students):
                return enrollment

    def generate_unique_staff_id(self):
        while True:
            staff_id = 'S' + str(random.randint(1000, 9999))
            if not any(staff['id'] == staff_id for staff in self.staffs):
                return staff_id

    def generate_unique_course_id(self):
        while True:
            course_id = 'C' + str(random.randint(1000, 9999))
            if not any(course['id'] == course_id for course in self.courses):
                return course_id

    # Email and phone validation constraints  
    def validate_email(self, email):
        return '@' in email and '.' in email and not any(
            student['email'] == email for student in self.students
        ) and not any(staff['email'] == email for staff in self.staffs)

    def validate_phone(self, phone):
        return len(phone) == 10 and phone.isdigit()

    # Student management methods
    def add_student(self):
        print("\nAdding New Student")
        print("-" * 20)

        name = input("Enter student name: ").strip()
        if not name:
            print("Error: Name is required!")
            return

        email = input("Enter student email: ").strip()
        if not email:
            print("Error: Email is required!")
            return
        if not self.validate_email(email):
            print("Error: Invalid or duplicate email!")
            return

        phone = input("Enter student phone: ").strip()
        if not self.validate_phone(phone):
            print("Error: Invalid phone number!")
            return

        address = input("Enter student address: ").strip()
        if not address:
            print("Error: Address is required!")
            return

        if not self.courses:
            print("Error: No courses available. Please add courses first.")
            return

        print("\nAvailable Courses:")
        for idx, course in enumerate(self.courses, 1):
            print(f"{idx}. {course['name']}")
        
        try:
            course_choice = int(input("Select course (enter number): "))
            if course_choice < 1 or course_choice > len(self.courses):
                print("Invalid course selection!")
                return
            course = self.courses[course_choice - 1]['name']
        except ValueError:
            print("Invalid input!")
            return

        student = {
            'id': self.generate_unique_student_id(),
            'enrollment': self.generate_enrollment_number(),
            'name': name,
            'email': email,
            'phone': phone,
            'address': address,
            'course': course
        }

        self.students.append(student)
        self.save_to_excel()
        print(f"\nStudent added successfully! ID: {student['id']}")

    def delete_student(self):
        student_id = input("Enter student ID to delete: ")
        try:
            student_id = int(student_id)
            student = next((s for s in self.students if s['id'] == student_id), None)
            if student:
                self.students.remove(student)
                self.save_to_excel()
                print("Student deleted successfully!")
            else:
                print("Student not found!")
        except ValueError:
            print("Invalid ID format!")

    def display_all_students(self):
        if not self.students:
            print("No students found!")
            return
        
        headers = ['ID', 'Enrollment', 'Name', 'Email', 'Phone', 'Address', 'Course']
        data = [[s['id'], s['enrollment'], s['name'], s['email'], s['phone'], s['address'], s['course']] 
                for s in self.students]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def display_student_by_id(self):
        student_id = input("Enter student ID: ")
        try:
            student_id = int(student_id)
            student = next((s for s in self.students if s['id'] == student_id), None)
            
            if student:
                self.display_student(student)
            else:
                print("Student not found!")
        except ValueError:
            print("Invalid ID format!")

    def display_student(self, student):
        print("\n----Student Details----")
        print(f"\nStudent ID: {student['id']}")
        print(f"Enrollment: {student['enrollment']}")
        print(f"Name: {student['name']}")
        print(f"Email: {student['email']}")
        print(f"Phone: {student['phone']}")
        print(f"Address: {student['address']}")
        print(f"Course: {student['course']}")

    def update_student(self):
        student_id = input("Enter student ID to update: ")
        try:
            student_id = int(student_id)
            student = next((s for s in self.students if s['id'] == student_id), None)
            if student:
                self.update_student_details(student)
            else:
                print("Student not found!")
        except ValueError:
            print("Invalid ID format!")

    def update_student_details(self, student):
        print("\nUpdating Student Details")
        print("-" * 20)

        email = input(f"Current email ({student['email']}): ").strip()
        if email and self.validate_email(email):
            student['email'] = email

        phone = input(f"Current phone ({student['phone']}): ").strip()
        if phone and self.validate_phone(phone):
            student['phone'] = phone

        address = input(f"Current address ({student['address']}): ").strip()
        if address:
            student['address'] = address

        self.save_to_excel()
        print("Student updated successfully!")

    # Staff management methods
    def add_staff(self):
        print("\nAdding New Staff")
        print("-" * 20)

        name = input("Enter staff name: ").strip()
        if not name:
            print("Error: Name is required!")
            return

        email = input("Enter staff email: ").strip()
        if not email:
            print("Error: Email is required!")
            return
        if not self.validate_email(email):
            print("Error: Invalid or duplicate email!")
            return

        phone = input("Enter staff phone: ").strip()
        if not self.validate_phone(phone):
            print("Error: Invalid phone number!")
            return

        if not self.courses:
            print("Error: No courses available. Please add courses first.")
            return

        print("\nAvailable Courses:")
        for idx, course in enumerate(self.courses, 1):
            print(f"{idx}. {course['name']}")
        
        try:
            course_choice = int(input("Select course (enter number): "))
            if course_choice < 1 or course_choice > len(self.courses):
                print("Invalid course selection!")
                return
            course = self.courses[course_choice - 1]['name']
        except ValueError:
            print("Invalid input!")
            return

        staff = {
            'id': self.generate_unique_staff_id(),
            'name': name,
            'email': email,
            'phone': phone,
            'course': course
        }

        self.staffs.append(staff)
        self.save_to_excel()
        print(f"\nStaff added successfully! ID: {staff['id']}")

    def delete_staff(self):
        staff_id = input("Enter staff ID to delete: ")
        staff = next((s for s in self.staffs if s['id'] == staff_id), None)
        if staff:
            self.staffs.remove(staff)
            self.save_to_excel()
            print("Staff deleted successfully!")
        else:
            print("Staff not found!")

    # Search for staff by ID
    def search_staff_by_id(self):
        staff_id = input("Enter staff ID to search: ").strip()
        staff = next((s for s in self.staffs if s['id'] == staff_id), None)
        if staff:
            self.display_staff(staff)
        else:
            print("Staff not found!")

    def display_staff(self, staff):
        print("\n----Staff Details----")
        print(f"\nStaff ID: {staff['id']}")
        print(f"Name: {staff['name']}")
        print(f"Email: {staff['email']}")
        print(f"Phone: {staff['phone']}")
        print(f"Course: {staff['course']}")


    def display_all_staffs(self):
        if not self.staffs:
            print("No staff found!")
            return
        
        headers = ['ID', 'Name', 'Email', 'Phone', 'Course']
        data = [[s['id'], s['name'], s['email'], s['phone'], s['course']] 
                for s in self.staffs]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def update_staff(self):
        staff_id = input("Enter staff ID to update: ")
        staff = next((s for s in self.staffs if s['id'] == staff_id), None)
        if staff:
            self.update_staff_details(staff)
        else:
            print("Staff not found!")

    def update_staff_details(self, staff):
        print("\nUpdating Staff Details")
        print("-" * 20)

        email = input(f"Current email ({staff['email']}): ").strip()
        if email and self.validate_email(email):
            staff['email'] = email

        phone = input(f"Current phone ({staff['phone']}): ").strip()
        if phone and self.validate_phone(phone):
            staff['phone'] = phone

        if self.courses:
            print("\nAvailable Courses:")
            for idx, course in enumerate(self.courses, 1):
                print(f"{idx}. {course['name']}")
            
            try:
                course_choice = int(input("Select new course (enter number): "))
                if course_choice < 1 or course_choice > len(self.courses):
                    print("Invalid course selection!")
                    return
                selected_course = self.courses[course_choice - 1]
                staff['course'] = selected_course['name']  # Update staff's assigned course

            # Update the assigned staff in the selected course
                selected_course['assigned_staff'] = staff['name']  # Set assigned staff to staff name
            except ValueError:
                print("Invalid input!")
                return

        self.save_to_excel()
        print("Staff updated successfully!")

    # Course management methods
    def add_course(self):
        print("\nAdding New Course")
        print("-" * 20)

        name = input("Enter course name: ").strip()
        if not name:
            print("Error: Course name is required!")
            return

        course = {
            'id': self.generate_unique_course_id(),
            'name': name,
            'assigned_staff': "None"
        }

        self.courses.append(course)
        self.save_to_excel()
        print(f"\nCourse added successfully! ID: {course['id']}")

    def delete_course(self):
        course_id = input("Enter course ID to delete: ")
        course = next((c for c in self.courses if c['id'] == course_id), None)
        if course:
            self.courses.remove(course)
            self.save_to_excel()
            print("Course deleted successfully!")
        else:
            print("Course not found!")

    # Search for course by ID
    def search_course_by_id(self):
        course_id = input("Enter course ID to search: ").strip()
        course = next((c for c in self.courses if c['id'] == course_id), None)
        if course:
            self.display_course(course)
        else:
            print("Course not found!")


    def display_course(self, course):
        print("\n----Course Details----")
        print(f"\nCourse ID: {course['id']}")
        print(f"Course Name: {course['name']}")
        print(f"Assigned Staff: {course['assigned_staff']}")


    def display_all_courses(self):
        if not self.courses:
            print("No courses found!")
            return
        
        headers = ['ID', 'Course Name', 'Assigned Staff']
        data = [[c['id'], c['name'], c['assigned_staff']] 
                for c in self.courses]
        print(tabulate(data, headers=headers, tablefmt='grid'))

    def update_course(self):
        course_id = input("Enter course ID to update: ")
        course = next((c for c in self.courses if c['id'] == course_id), None)
        if course:
            self.update_course_details(course)
        else:
            print("Course not found!")

    def update_course_details(self, course):
        print("\nUpdating Course Details")
        print("-" * 20)

        if self.staffs:
            print("\nAvailable Staff:")
            for idx, staff in enumerate(self.staffs, 1):
                print(f"{idx}. {staff['name']}")
            
            try:
                staff_choice = int(input("Select staff (enter number): "))
                if staff_choice < 1 or staff_choice > len(self.staffs):
                    print("Invalid staff selection!")
                    return
                staff = self.staffs[staff_choice - 1]
                course['assigned_staff'] = staff['name']
                staff['course'] = course['name']  # Update staff's assigned course
            except ValueError:
                print("Invalid input!")
                return

        self.save_to_excel()
        print("Course updated successfully!")

        
def main():
    system = StudentManagementSystem()

    while True:
        print("\n-------- School Management System --------")
        print("1. Student Options")
        print("2. Staff Options")
        print("3. Course Options")
        print("4. Exit")

        choice = input("Enter your choice: ").strip()
        
        if choice == '1':
            student_options(system)
        elif choice == '2':
            staff_options(system)
        elif choice == '3':
            course_options(system)
        elif choice == '4':
            print("Exiting the system...")
            break
        else:
            print("Invalid choice! Please select again.")

def student_options(system):
    while True:
        print("\n-------- Student Options --------")
        print("a. Add Student")
        print("b. Delete Student by ID")
        print("c. Display All Students")
        print("d. Search Student by ID")
        print("e. Update Student by ID")
        print("f. Back to Main Menu")
        
        choice = input("Enter your choice: ").strip()

        if choice == 'a':
            system.add_student()
        elif choice == 'b':
            system.delete_student()
        elif choice == 'c':
            system.display_all_students()
        elif choice == 'd':
            system.display_student_by_id()
        elif choice == 'e':
            system.update_student()
        elif choice == 'f':
            break
        else:
            print("Invalid choice! Please select again.")

def staff_options(system):
    while True:
        print("\n-------- Staff Options --------")
        print("a. Add Staff")
        print("b. Delete Staff by ID")
        print("c. Display All Staffs")
        print("d. Update Staff by ID")
        print("e. Search Staff by ID")
        print("f. Back to Main Menu")

        choice = input("Enter your choice: ").strip()

        if choice == 'a':
            system.add_staff()
        elif choice == 'b':
            system.delete_staff()
        elif choice == 'c':
            system.display_all_staffs()
        elif choice == 'd':
            system.update_staff()
        elif choice == 'e':
            system.search_staff_by_id()
        elif choice == 'f':
            break
        else:
            print("Invalid choice! Please select again.")

def course_options(system):
    while True:
        print("\n-------- Course Options --------")
        print("a. Add Course")
        print("b. Delete Course by ID")
        print("c. Display All Courses")
        print("d. Update Course by ID")
        print("e. Search Course by ID")
        print("f. Back to Main Menu")

        choice = input("Enter your choice: ").strip()

        if choice == 'a':
            system.add_course()
        elif choice == 'b':
            system.delete_course()
        elif choice == 'c':
            system.display_all_courses()
        elif choice == 'd':
            system.update_course()
        elif choice == 'e':
            system.search_course_by_id()
        elif choice == 'f':
            break
        else:
            print("Invalid choice! Please select again.")

if __name__ == "__main__":
    main()

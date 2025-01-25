#beki
from openpyxl import Workbook, load_workbook
import os,datetime

TODAY = datetime.date.today()

home_dir = os.path.expanduser('~')

bk_xlsx = os.path.join(home_dir, 'Desktop','bk','bk.xlsx')

UKLSalesSummaryT= os.path.join(home_dir, 'Desktop','bk','UKL Sales Summary(T).xlsx')

UKLSalesSummaryd = os.path.join(home_dir, 'Desktop','bk','UKL Sales Summary(d).xlsx')

UKLPhysicalStockReports = os.path.join(home_dir, 'Desktop','bk','UKL Physical Stock Report(s).xlsx')

UKLSalesSummaryrt = os.path.join(home_dir, 'Desktop','bk','UKL Sales Summary(rt).xlsx')

wb_mtd = load_workbook(UKLSalesSummaryT,data_only = True)
ws_mtd = wb_mtd['UKL Sales Summary']

wb_retail = load_workbook(UKLSalesSummaryrt,data_only = True)
ws_retail = wb_retail['UKL Sales Summary']

wb_daily = load_workbook(UKLSalesSummaryd,data_only = True)
ws_daily = wb_daily['UKL Sales Summary']

wb_stock = load_workbook(UKLPhysicalStockReports,data_only = True)
ws_stock = wb_stock['UKL Physical Stock Report']

wb_bk = load_workbook(bk_xlsx)
ws_main = wb_bk['MAIN']

ws_report = wb_bk['REPORT']
ws_rtcont = wb_bk['RTcont']

ws_report['B2'] = TODAY





#stock section
stock_mapping = {
    'KNORR ALL IN ONE CUBE 240X8G': 'f16',
    'LIFEBUOY SOAP TOTAL 12X6X70G': 'f17',
    'OMO HW POWDER GAIA 72X100G': 'f18',
    'SUNLIGHT HW POWDER 72X90G': 'f19',
    'SUNLIGHT WHITE UNWRAPED BAR SOAP 50X200G': 'f20',
    'OMO HW POWDER GAIA 100X40G': 'f21',
    'SUNLIGHT HW POWDER 100X40G': 'f22',
    'LUX SOAP SOFT TOUCH 12X6X70G': None,
    'LUX SOAP SOFT CARESS 12X6X70G': None, 
    'SUNSILK CONDITIONER AVOCADO 12X350ML': None,  
    'SUNSILK CONDITIONER COCONUT 12X350ML': None, 
    'SUNSILK SHAMPOO COCONUT 12X350ML': None,  
    'SUNSILK SHAMPOO AVOCADO 12X350ML': None, 
    'SUNSILK SHAMPOO COCONUT 12X700ML': None,  
    'SUNSILK CONDITIONER COCONUT 12X700ML': None,  
    'SUNSILK SHAMPOO AVOCADO 12X700ML': None,  
    'SUNSILK CONDITIONER AVOCADO 12X700ML': None, 
    'LIFEBUOY SOAP BAR LEMON 12X6X70G': 'f26'
}


lux_sum = 0
sunsilk_350_sum = 0
sunsilk_700_sum = 0


for i in range(11, 300):
    stock = ws_stock[f'd{i}'].value  
    value = ws_stock[f'j{i}'].value  
    
    if stock and stock in stock_mapping:
        
        if stock in ['LUX SOAP SOFT TOUCH 12X6X70G', 'LUX SOAP SOFT CARESS 12X6X70G']:
            lux_sum += value if value else 0
        
     
        elif stock in [
            'SUNSILK CONDITIONER AVOCADO 12X350ML', 
            'SUNSILK CONDITIONER COCONUT 12X350ML', 
            'SUNSILK SHAMPOO COCONUT 12X350ML', 
            'SUNSILK SHAMPOO AVOCADO 12X350ML'
        ]:
            sunsilk_350_sum += value if value else 0

        elif stock in [
            'SUNSILK SHAMPOO COCONUT 12X700ML',
            'SUNSILK CONDITIONER COCONUT 12X700ML',
            'SUNSILK SHAMPOO AVOCADO 12X700ML',
            'SUNSILK CONDITIONER AVOCADO 12X700ML'
        ]:
            sunsilk_700_sum += value if value else 0
        elif stock_mapping[stock]:
            cell = stock_mapping[stock]
            ws_report[cell] = value


ws_report['f25'] = lux_sum 
ws_report['f23'] = sunsilk_350_sum 
ws_report['f24'] = sunsilk_700_sum  

#stock value				
for i in range(11,1000):
		total=ws_stock['j'+str(i)].value
		value=ws_stock['w'+str(i)].value
		
		if not total== None:
				if total== 'Grand Total':
					ws_report['F13'] = value
					break

#‐-------stock section end---------



# Mapping for MTD and retail
product_mapping = {
    'KNORR BOUILLON CUBES BEEF 8G': 'B10',
    'LIFEBUOY PW BAR SOAP TOTAL 70G': 'B11',
    'SIGNAL TOTHPASTE ESS CAVITY FIGHTER 60G': 'B12',
    'OMO NM STD POWDER 100G': 'B13',
    'SUNLIGHT NM STD HW POWDER 90G': 'B14',
    'SUNLIGHT DA HARD SOAP WHITE 200G': 'B15',
    'KNORR BOUILLONS CHICKEN CUBES 8G': 'B16',
    'LUX SKIN CLEANSING BAR SOFT CARESS 70G': 'B17',
    'LUX SKIN CLEANSING BAR SOFT TOUCH 70G': 'B18',
    'OMO NM STD POWDER HW GAIA 160G': 'B19',
    'SUNLIGHT NM STD HW POWDER 160G': 'B20',
    'SUNSILK REG RINSE OUT COND AVOCADO 350ML': 'B21',
    'SUNSILK SHAMPOO AVOCADO 350ML': 'B22',
    'SUNSILK SHAMPOO COCONUT 350ML': 'B23',
    'SUNSILK REG RINSE OUT COND COCONUT 350ML': 'B24',
    'LIFEBUOY PW BAR LEMON FRESH 70G': 'B25',
    'OMO NM STD HW POWDER GAIA 1KG': 'B26',
    'SUNLIGHT NM STD HW POWDER 40G': 'B27',
    'SIGNAL TOTHPASTE ESS CAVITY FIGHTER 140G': 'B28',
    'OMO NM STD HW POWDER GAIA 40G': 'B29'
}

# Initialize all product mapping cells to 0
for product, cell in product_mapping.items():
    ws_main[cell] = 0  # Default to 0 in ws_main
    base_cell = cell
    row_number = int(base_cell[1:]) - 8  # Adjust row number as needed
    ws_rtcont[base_cell[0] + str(row_number)] = 0
    

    
    



# Iterate through rows in the Excel sheets
for i in range(17, 300):
    # Read values from MTD sheet
    mtd_sku = ws_mtd['y' + str(i)].value
    mtd_value = ws_mtd['ao' + str(i)].value
    mtd_value = int(mtd_value) if mtd_value is not None else 0  # Default to 0 if None

    # Read values from Retail sheet
    retail_sku = ws_retail['y' + str(i)].value
    retail_value = ws_retail['ao' + str(i)].value
    retail_value = int(retail_value) if retail_value is not None else 0  


    # MTD data
    if mtd_sku and mtd_sku in product_mapping:
        ws_main[product_mapping[mtd_sku]] = mtd_value

    # Retail data
    if retail_sku and retail_sku in product_mapping:
        base_cell = product_mapping[retail_sku]
        
        row_number = int(base_cell[1:]) - 8  
        ws_rtcont[base_cell[0] + str(row_number)] = retail_value
        

        
    elif not retail_sku and not mtd_sku:
        break  



#daily section

daily_mapping = {
'KNORR BOUILLON CUBES BEEF 8G': 'B16',
        'LIFEBUOY PW BAR SOAP TOTAL 70G': 'B17',
        'OMO NM STD POWDER 100G': 'B18',
        'SUNLIGHT NM STD HW POWDER 90G': 'B19',
        'SUNLIGHT DA HARD SOAP WHITE 200G': 'B20',
         'OMO NM STD HW POWDER GAIA 40G': 'B21',
          'SUNLIGHT NM STD HW POWDER 40G': 'B22',
        'SUNSILK REG RINSE OUT COND AVOCADO 350ML': None,
        'SUNSILK SHAMPOO AVOCADO 350ML': None,
        'SUNSILK SHAMPOO COCONUT 350ML': None,
        'SUNSILK REG RINSE OUT COND COCONUT 350ML': None,
        'SUNSILK REG RINSE OUT COND AVOCADO 700ML': None,
        'SUNSILK SHAMPOO AVOCADO 700ML': None,
        'SUNSILK SHAMPOO COCONUT 700ML': None,
        'SUNSILK REG RINSE OUT COND COCONUT 700ML': None,
        'LUX SKIN CLEANSING BAR SOFT CARESS 70G':None,
        'LUX SKIN CLEANSING BAR SOFT TOUCH 70G': None,
        'LIFEBUOY PW BAR LEMON FRESH 70G': 'B26'
       
        }
#daily sales        
for product, cell in daily_mapping.items():
    if cell:
        ws_report[cell] = 0
        
lux_sum_d = 0
sunsilk_350_sum_d = 0
sunsilk_700_sum_d = 0

for i in range(11, 300):
    daily = ws_daily[f'y{i}'].value  
    value = ws_daily[f'ao{i}'].value
    #value = int(value) if value and isinstance(value, (int, float)) else 0

    
    if daily and daily in daily_mapping:
        
        if daily in ['LUX SKIN CLEANSING BAR SOFT CARESS 70G',
        'LUX SKIN CLEANSING BAR SOFT TOUCH 70G']:
            lux_sum_d += value if value else 0
        
     
        elif daily in [
            'SUNSILK REG RINSE OUT COND AVOCADO 350ML',
        'SUNSILK SHAMPOO AVOCADO 350ML',
        'SUNSILK SHAMPOO COCONUT 350ML',
        'SUNSILK REG RINSE OUT COND COCONUT 350ML'
        ]:
            sunsilk_350_sum_d += value if value else 0

        elif daily in [
            'SUNSILK REG RINSE OUT COND AVOCADO 700ML',
        'SUNSILK SHAMPOO AVOCADO 700ML',
        'SUNSILK SHAMPOO COCONUT 700ML',
        'SUNSILK REG RINSE OUT COND COCONUT 700ML'
        ]:
            sunsilk_700_sum_d += value if value else 0
        elif daily_mapping[daily]:
            cell = daily_mapping[daily]
            ws_report[cell] = value


ws_report['b25'] = lux_sum_d 
ws_report['b23'] = sunsilk_350_sum_d 
ws_report['b24'] = sunsilk_700_sum_d  




#daily total sales
for i in range(11,300):
				amount = ws_daily['ai'+str(i)].value
				if amount == 'Net Amount':
					i+=2
					ws_report['B13'] = ws_daily['ai'+str(i)].value


#daily distribution

for product, cell in daily_mapping.items():
    if cell:
        base_cell = cell
        main_cell = int(base_cell[1:]) + 13
        dist_cell = base_cell[0]+ str(main_cell)
        ws_report[dist_cell] = 0
        
lux_sum = 0
sunsilk_350_sum = 0
sunsilk_700_sum = 0
ss_700_sum = 0


for i in range(11, 300):
    daily_dist = ws_daily[f'Y{i}'].value  
    value_dist = ws_daily[f'AJ{i}'].value
    
    mtd_ss = ws_mtd[f'Y{i}'].value  
    value_ss = ws_mtd[f'AO{i}'].value


    
    if daily_dist and daily_dist in daily_mapping:
        
        
        if daily_dist in ['LUX SKIN CLEANSING BAR SOFT CARESS 70G',
        'LUX SKIN CLEANSING BAR SOFT TOUCH 70G']:
            lux_sum += value_dist if value_dist else 0
        
     
        elif daily_dist in [
            'SUNSILK REG RINSE OUT COND AVOCADO 350ML',
        'SUNSILK SHAMPOO AVOCADO 350ML',
        'SUNSILK SHAMPOO COCONUT 350ML',
        'SUNSILK REG RINSE OUT COND COCONUT 350ML'
        ]:
            sunsilk_350_sum += value_dist if value_dist else 0

        elif daily_dist in [
            'SUNSILK REG RINSE OUT COND AVOCADO 700ML',
        'SUNSILK SHAMPOO AVOCADO 700ML',
        'SUNSILK SHAMPOO COCONUT 700ML',
        'SUNSILK REG RINSE OUT COND COCONUT 700ML'
        ]:
            sunsilk_700_sum += value_dist if value_dist else 0
        elif daily_mapping[daily_dist]:
            cell = daily_mapping[daily_dist]
            base_cell = cell
            main_cell = int(base_cell[1:]) + 13
            dist_cell = base_cell[0]+ str(main_cell)
            ws_report[dist_cell] = value_dist

#mtd ss700ml       
    if mtd_ss and mtd_ss in daily_mapping:       
        if mtd_ss in [
            'SUNSILK REG RINSE OUT COND AVOCADO 700ML',
        'SUNSILK SHAMPOO AVOCADO 700ML',
        'SUNSILK SHAMPOO COCONUT 700ML',
        'SUNSILK REG RINSE OUT COND COCONUT 700ML'
        ]:
            ss_700_sum += value_ss if value_ss else 0


ws_report['C24'] = ss_700_sum 
#---------mtd ss700ml end---


ws_report['b38'] = lux_sum 
ws_report['b36'] = sunsilk_350_sum 
ws_report['b37'] = sunsilk_700_sum 

					
#mtd distribution

for product, cell in daily_mapping.items():
    if cell:
        base_cell = cell
        main_cell = int(base_cell[1:]) + 13
        dist_cell = 'C' + str(main_cell)
        ws_report[dist_cell] = 0
        
lux_sum = 0
sunsilk_350_sum = 0
sunsilk_700_sum = 0


for i in range(11, 300):
    mtd_dist = ws_mtd[f'Y{i}'].value  
    value_dist = ws_mtd[f'AJ{i}'].value


    
    if mtd_dist and mtd_dist in daily_mapping:
        
        
        if mtd_dist in ['LUX SKIN CLEANSING BAR SOFT CARESS 70G',
        'LUX SKIN CLEANSING BAR SOFT TOUCH 70G']:
            lux_sum += value_dist if value_dist else 0
        
     
        elif mtd_dist in [
            'SUNSILK REG RINSE OUT COND AVOCADO 350ML',
        'SUNSILK SHAMPOO AVOCADO 350ML',
        'SUNSILK SHAMPOO COCONUT 350ML',
        'SUNSILK REG RINSE OUT COND COCONUT 350ML'
        ]:
            sunsilk_350_sum += value_dist if value_dist else 0

        elif mtd_dist in [
            'SUNSILK REG RINSE OUT COND AVOCADO 700ML',
        'SUNSILK SHAMPOO AVOCADO 700ML',
        'SUNSILK SHAMPOO COCONUT 700ML',
        'SUNSILK REG RINSE OUT COND COCONUT 700ML'
        ]:
            sunsilk_700_sum += value_dist if value_dist else 0
        elif daily_mapping[mtd_dist]:
            cell = daily_mapping[mtd_dist]
            base_cell = cell
            main_cell = int(base_cell[1:]) + 13
            dist_cell = 'C' + str(main_cell)
            ws_report[dist_cell] = value_dist
            


ws_report['C38'] = lux_sum 
ws_report['C36'] = sunsilk_350_sum 
ws_report['C37'] = sunsilk_700_sum  					
					

				
												
										
					
#mtd total sales			
for i in range(11,300):
				amount = ws_mtd['ai'+str(i)].value
				if amount == 'Net Amount':
					i+=2
					ws_main['I10'] = ws_mtd['ai'+str(i)].value

#retail total sales
for i in range(11,300):
				amount = ws_retail['ai'+str(i)].value
				if amount == 'Net Amount':
					i+=2
					ws_main['I11'] = ws_retail['ai'+str(i)].value

def save():
	wb_bk.save(bk_xlsx)
	

if __name__ == '__main__':
	save()

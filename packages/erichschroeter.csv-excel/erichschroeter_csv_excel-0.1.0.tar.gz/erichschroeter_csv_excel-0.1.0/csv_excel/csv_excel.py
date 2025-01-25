import csv
import importlib
import logging
import openpyxl
import os
from pathlib import Path
import xlsxwriter
from xlsxwriter.utility import xl_cell_to_rowcol
import yaml


def column_to_index(col_str):
    """
    Convert a column cell reference notation to a zero indexed row and column.
    For example, 'A' will assume 'A1' and return (0,0)

    Args:
       col_str:  The column for A1 style string.

    Returns:
        row, col: Zero indexed cell row and column indices.

    """
    return xl_cell_to_rowcol(f"{col_str.upper()}1")[1]


class RuleError(Exception):
    def __init__(self, rule_file, message, *args: object) -> None:
        self.message = f"{os.path.basename(rule_file)}: {message}"
        super().__init__(message, *args)


class CsvRuleError(RuleError):
    def __init__(self, rule_file, row, col, message, *args: object) -> None:
        super().__init__(rule_file, f"(row: {row}, col: {col}): {message}", *args)


class WorkbookFactory:
    def __init__(self, config_path=None) -> None:
        self.config_path = config_path
        self.config = None
        if self.config_path:
            with open(self.config_path, "r") as yamlfile:
                self.config = yaml.safe_load(yamlfile)

    def _csv_path_to_worksheet_title(self, csv_path) -> str:
        title = os.path.basename(csv_path)  # don't include full path, just file name
        title = os.path.splitext(title)[0]  # remove extension
        return title

    def build_openpyxl(self, csv_files, output_path=None):
        wb = openpyxl.Workbook()
        # Delete the default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        for csv_file in csv_files:
            with open(csv_file, "r") as f:
                reader = csv.reader(f)
                csv_data = list(reader)

            worksheet_title = self._csv_path_to_worksheet_title(csv_file)
            sheet = wb.create_sheet(title=worksheet_title)
            logging.debug(f'Added worksheet "{worksheet_title}"')

            # Write the data to the worksheet
            for data in csv_data:
                sheet.append(data)
        if output_path:
            wb.save(output_path)
        return wb

    def build_xlsxwriter(self, csv_files, output_path):
        wb = xlsxwriter.Workbook(output_path)
        # Delete the default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        # Include the Excel macro that auto exports worksheets to CSV files when file is saved.
        vbaproject_path = f"{os.path.dirname(os.path.abspath(__file__))}/vbaProject.bin"
        logging.debug(f'Packing VBA project into Excel file: "{vbaproject_path}"')
        wb.add_vba_project(vbaproject_path)

        for csv_file in csv_files:
            with open(csv_file, "r") as f:
                reader = csv.reader(f)
                csv_data = list(reader)

            worksheet_title = self._csv_path_to_worksheet_title(csv_file)
            sheet = wb.add_worksheet(name=worksheet_title)
            logging.debug(f'Added worksheet "{worksheet_title}"')

            # Apply any config specifications.
            if self.config:
                if worksheet_title in self.config["sheets"]:
                    sheet_config = self.config["sheets"][worksheet_title]
                    if "columns" in sheet_config:
                        for colname, colcfg in sheet_config["columns"].items():
                            if "width" in colcfg:
                                width = int(colcfg["width"])
                                colindex = column_to_index(colname)
                                sheet.set_column_pixels(colindex, colindex, width)
                                logging.debug(
                                    f'Sheet "{worksheet_title}" column "{colname}" ({colindex}) to {width}px'
                                )

            # Write the data to the worksheet
            for row, data in enumerate(csv_data):
                sheet.write_row(row, 0, data)

        return wb


def csv2xl(args):
    """
    Generates or updates an Excel file from multiple CSV files.

    Args:
        args:  The command line args.
    """
    # Use xlsxwriter due to support for vbaProject macros.
    wb = WorkbookFactory(args.config).build_xlsxwriter(args.csv_files, args.output)
    # Save the workbook
    wb.close()


def xl2csv(args):
    """
    Exports worksheets within an Excel file to CSV files.

    Args:
        args:  The command line args.
    """
    wb = openpyxl.load_workbook(args.file)
    # Create the output directory if it does not exist.
    if args.output_dir:
        Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    for sheet in wb:
        with open(
            os.path.join(
                args.output_dir if args.output_dir else "", f"{sheet.title}.csv"
            ),
            "w+",
            newline="",
        ) as f:
            logging.debug(f'Exporting worksheet "{sheet.title}"')
            c = csv.writer(f)
            for row in sheet.rows:
                c.writerow([cell.value for cell in row])


def validate(args):
    # Use openpyxl due to better support for reading data.
    wb = WorkbookFactory(args.config).build_openpyxl(args.csv_files)

    from os.path import dirname, basename, isfile, join
    import glob

    modules = glob.glob(join(args.rules_dir, "*.py"))
    logging.debug(f"Found modules: {modules}")
    # Use the rules_dir as a python module and import all .py files, excluding __init__.py
    rule_modules = [
        f"{basename(dirname(args.rules_dir))}.{basename(f)[:-3]}"
        for f in modules
        if isfile(f) and not f.endswith("__init__.py")
    ]
    logging.debug(f"Checking rules: {rule_modules}")
    results = []
    for rule_module in rule_modules:
        rule = importlib.import_module(rule_module)
        v = getattr(rule, "validate")
        result = v(wb)
        if result:
            results.extend(result)
    if results:
        for result in results:
            logging.error(f"{result.message}")

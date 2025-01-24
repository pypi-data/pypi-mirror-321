import argparse
import sys

import click
import openpyxl
from Bio.Seq import Seq
from openpyxl.styles import Border, Color, Font, PatternFill, colors

from BASE_dzne.core.legacy.libBASE.core.varia import (
    AlignPCRObject,
    SequenceFile,
    exportDict,
    updateExcelRow,
)

__all__ = ["main", "run"]

REDFILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
ORANGEFILL = PatternFill(start_color="d1731e", end_color="d1731e", fill_type="solid")
LIGHTGREENFILL = PatternFill(
    start_color="92D050", end_color="92D050", fill_type="solid"
)
DEEPGREENFILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
PURPLEFILL = PatternFill(start_color="826aaf", end_color="826aaf", fill_type="solid")
NOFILL = PatternFill(fill_type=None)


def getfill(mark):
    if mark == "O":
        return ORANGEFILL
    if mark == "R":
        return REDFILL
    if mark == "":
        return NOFILL
    if mark == "P":
        return PURPLEFILL
    if mark == "D":
        return DEEPGREENFILL
    if mark == "L":
        return LIGHTGREENFILL
    raise ValueError


def getgreen(mark):
    return mark in ("L", "D")


def getmark(output):
    if output.find("WARNING") != -1:
        return "O"
    elif output.find("nsSHM+ fr1") != -1:
        if (
            output[output.find("nsSHM+ fr1") + 1 :].find("nsSHM+") != -1
        ):  # if we have additional nsSHM in other regions than fr1, we mark the cell as red
            return "R"
        else:  # else it's only orange (maybe we overlooked something about the primer?)
            return "O"
    elif output.find("nsSHM+") != -1:
        return "R"
    elif output.find("do not match") != -1:
        return "R"
    elif output.find("index") != -1:
        return "R"
    elif output.find("Uncaught") != -1:
        return "R"
    elif output.find("FileNotFound") != -1:
        return ""
    elif output.find("BQ") != -1:
        return "R"
    elif output.find("nsSHM-") != -1:
        return "O"
    elif output.find("differ") != -1:
        return "O"
    elif output.find("ikely") != -1:
        return "O"
    elif output.find("empty") != -1:
        return "R"
    elif output.find("not productive") != -1:
        return "P"
    elif output == "0":
        return "D"
    else:
        return "L"


@click.command(name="cBASE")
@click.argument("infile")
@click.argument("outfile")
@click.option(
    "--dataprefix",
    help="Prefix for the sequence file filenames. This can be a directory name.",
)
@click.option(
    "--pcr2read",
    help="Column where the name of the sequencing file of the 2nd PCR is found.",
)
@click.option(
    "--plasmidread",
    help="Column where the name of the sequencing file of the plasmid is found.",
)
@click.option(
    "--shmanalysis",
    help=(
        "Columns where the somatic hypermutation analysis should be written to. "
        "If four columns (separated by a comma) instead of one are given, the software will also "
        "give a more detailed analysis of the somatic hypermutations of the PCR2 read, "
        "of the plasmid, and the idealized antibody."
    ),
)
def main(**kwargs):
    """
    cBase compares the sequencing data of plasmids and PCR reads on a nucleotide per nucleotide basis.
    """
    run(**kwargs)


def run(
    infile,
    outfile,
    *,
    dataprefix=None,
    pcr2read=None,
    plasmidread=None,
    shmanalysis=None,
):
    nsobj = argparse.Namespace()
    nsobj.dataprefix = dataprefix
    nsobj.pcr2read = pcr2read
    nsobj.plasmidread = plasmidread
    nsobj.shmanalysis = shmanalysis

    if infile == outfile:
        sys.exit("infile is outfile. Please do not do that. Aborting ...")

    try:
        workbook = openpyxl.load_workbook(infile)
    except FileNotFoundError:
        sys.exit("File %s not found! Aborting..." % infile)
    except ValueError as my_err:
        sys.exit(
            "OOPS! An error occured while parsing "
            + infile
            + ". "
            + str(my_err)
            + ". Aborting ..."
        )
    except OSError as my_err:
        sys.exit(
            "OOPS! An error occured while parsing "
            + infile
            + ". "
            + str(my_err)
            + ". Maybe the wrong filetype? Aborting ..."
        )

    ws = workbook.active

    if nsobj.pcr2read is not None:
        if nsobj.pcr2read.find(":") == -1:  # we suppose this is a single cell then
            nsobj.pcr2read = nsobj.pcr2read + ":" + nsobj.pcr2read
        pcr2read = workbook.active[nsobj.pcr2read]
    if nsobj.shmanalysis is not None:
        differential_analysis_column = nsobj.shmanalysis[0]
        if len(nsobj.shmanalysis) > 1:
            pcr2_shm_column = nsobj.shmanalysis.split(",")[1]
            plasmid_shm_column = nsobj.shmanalysis.split(",")[2]
            ideal_shm_column = nsobj.shmanalysis.split(",")[3]

    to_compare = []

    ###chains['H']] is loaded by chains['H']=workbook.active[nsobj.heavy]. if nsobj.heavy=Z (a whole row),
    ###a list is returned, but if nsobj.heavy=Z4:Z240 (for example..), then a tuple is returen (?!?).
    ### thats why we have to iterate over seq, instead of seq###
    output = ""
    green = False
    for (seq,) in pcr2read:
        ###TODO: error handling!
        if seq.value is None:
            continue
        filename_pcr2 = seq.value
        filename_pcr2 = nsobj.dataprefix + str(filename_pcr2) + ".ab1"

        # the following line is a workaround for the inconsistent naming scheme of Eurofins
        filename_pcr2 = filename_pcr2.replace("-", "_")

        try:
            pcr2 = SequenceFile(filename_pcr2)

            filename_plasmid = ws[nsobj.plasmidread + str(seq.row)].value
            filename_plasmid = nsobj.dataprefix + str(filename_plasmid) + ".ab1"

            # the following line is a workaround for the inconsistent naming scheme of Eurofins
            filename_plasmid = filename_plasmid.replace("-", "_")

            print("Comparing: " + filename_pcr2 + " " + filename_plasmid)

            try:
                plasmid = SequenceFile(filename_plasmid)
                if plasmid.successfullyParsed == True:
                    temp = AlignPCRObject(pcr2, plasmid)
                    output = temp.output
                else:
                    output = (
                        "BQ: Could not blast "
                        + plasmid.filename
                        + ". This could either be due to bad sequencing quality, or maybe because it's an empty vector? igblast complained: "
                        + plasmid.comment
                    )

            except FileNotFoundError as err:
                output = "FileNotFound"
                print(str(err))
            # except:
            #    output="Uncaught exception. Please inform the author of this software and provide the ab1-file(s)."

            output_cell = differential_analysis_column + str(seq.row)

            mark = getmark(output)
            fill = getfill(mark)
            green = getgreen(mark)
            ws[output_cell].fill = fill
            ws[output_cell] = output

            if len(nsobj.shmanalysis) > 1:
                try:
                    if temp.pcr1.successfullyParsed == True:
                        ed1 = exportDict(temp.pcr1)
                        ws[pcr2_shm_column + str(seq.row)] = str(ed1["SHM"])
                    else:
                        ws[pcr2_shm_column + str(seq.row)] = "n/a"
                except:
                    ws[pcr2_shm_column + str(seq.row)] = "n/a"
                try:
                    if temp.pcr2.successfullyParsed == True:
                        ed2 = exportDict(temp.pcr2)
                        ws[plasmid_shm_column + str(seq.row)] = str(ed2["SHM"])
                    else:
                        ws[plasmid_shm_column + str(seq.row)] = "n/a"
                except:
                    ws[plasmid_shm_column + str(seq.row)] = "n/a"
                try:
                    if (
                        temp.pcr2.successfullyParsed == True
                        and temp.pcr1.successfullyParsed == True
                    ):
                        ws[ideal_shm_column + str(seq.row)] = (
                            temp.number_of_shm_v_gene_ideal
                        )
                    else:
                        ws[ideal_shm_column + str(seq.row)] = "n/a"
                except:
                    ws[ideal_shm_column + str(seq.row)] = "n/a"

            if green is not True:
                print(output)

            output = ""
            temp = None
            green = False
        except FileNotFoundError as err:
            print(str(err))
        except ValueError as my_err:
            sys.exit(
                "OOPS! An error occured while parsing "
                + filename_pcr2
                + ". "
                + str(my_err)
                + ". Aborting ..."
            )
        except OSError as my_err:
            sys.exit(
                "OOPS! An error occured while parsing "
                + filename_pcr2
                + ". "
                + str(my_err)
                + ". Maybe the wrong filetype? Aborting ..."
            )

    workbook.save(outfile)

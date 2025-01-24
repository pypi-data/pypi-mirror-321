#!/opt/conda/bin/python3
import argparse

# importing os module
import os

# importing shutil module
import shutil
import sys

import click
import openpyxl
from openpyxl.utils.cell import get_column_letter

from BASE_dzne.core.legacy.libBASE.core.varia import SequenceFile, updateExcelRow


###helper
def createExportDict(ws, begin, end):
    """returns a dict with keys='what should go in this column' and values='excel column name', created from worksheet ws"""
    output_dict = dict()
    for row in ws[begin:end]:
        for cell in row:
            # 190328: this used to be
            # output_dict[cell.value]=cell.column
            # apparently now we have to use get_column_letter to convert the number into a letter
            output_dict[cell.value] = get_column_letter(cell.column)
    return output_dict


@click.command(name="aBASE")
@click.argument("infile")
@click.argument("outfile")
@click.option(
    "--dataprefix",
    help="Prefix for the sequence file filenames. This can be a directory name.",
)
@click.option(
    "--hchain",
    help="Column where the heavy chain is found.",
)
@click.option(
    "--kchain",
    help="Column where the kappa chain is found.",
)
@click.option(
    "--lchain",
    help="Column where the lambda chain is found.",
)
@click.option(
    "--heavykeys",
    help="Columns where the heavy chain output should be written to.",
)
@click.option(
    "--kappakeys",
    help="Columns where the kappa chain output should be written to.",
)
@click.option(
    "--lambdakeys",
    help="Columns where the lambda chain output should be written to.",
)
@click.option(
    "--overwrite",
    is_flag=True,
    help="Overwrite the existing output file.",
)
def main(**kwargs):
    """
    aBase automatically igblasts sequencing reads and gives cloning suggestions and immunological annotations.
    """
    run(**kwargs)


def run(
    infile,
    outfile,
    *,
    dataprefix=None,
    hchain=None,
    kchain=None,
    lchain=None,
    heavykeys=None,
    kappakeys=None,
    lambdakeys=None,
    overwrite=False,
):

    nsobj = argparse.Namespace()
    nsobj.infile = infile
    nsobj.outfile = outfile
    nsobj.dataprefix = dataprefix
    nsobj.hchain = hchain
    nsobj.kchain = kchain
    nsobj.lchain = lchain
    nsobj.heavykeys = heavykeys
    nsobj.kappakeys = kappakeys
    nsobj.lambdakeys = lambdakeys
    nsobj.overwrite = overwrite
    run_namespace(nsobj)


def run_namespace(nsobj, /):
    if nsobj.infile == nsobj.outfile:
        sys.exit("infile is outfile. Please do not do that. Aborting ...")

    try:
        # 2019-03-28: keep_vba=True was added since the file could not be saved else (see: https://bitbucket.org/openpyxl/openpyxl/issues/766/workbook-cannot-saved-twice)
        workbook = openpyxl.load_workbook(nsobj.infile, keep_vba=True)
    except FileNotFoundError:
        sys.exit("File " + nsobj.infile + " not found! Aborting...")
    except ValueError as my_err:
        sys.exit(
            "OOPS! An error occured while parsing "
            + nsobj.infile
            + ". "
            + str(my_err)
            + ". Aborting ..."
        )
    except OSError as my_err:
        sys.exit(
            "OOPS! An error occured while parsing "
            + nsobj.infile
            + ". "
            + str(my_err)
            + ". Maybe the wrong filetype? Aborting ..."
        )

    ws = workbook.active

    chains = {}
    begin = {}
    end = {}
    columndict = {}

    if nsobj.hchain is not None:

        if nsobj.hchain.find(":") == -1:  # we suppose this a single cell then
            nsobj.hchain = nsobj.hchain + ":" + nsobj.hchain

        chains["H"] = workbook.active[nsobj.hchain]
        if nsobj.heavykeys is not None:
            begin = nsobj.heavykeys.split(":")[0]
            end = nsobj.heavykeys.split(":")[1]
            columndict["H"] = createExportDict(ws, begin, end)
            if "Comment" not in columndict["H"].keys():
                sys.exit(
                    "No column titled 'Comment' in the fields specified by --heavykeys. Exiting..."
                )

        else:
            print("--heavykeys not set. No heavy chain data will be written")

    if nsobj.lchain is not None:
        if (
            nsobj.lchain.find(":") == -1
        ):  # we suppose this a single cell then. The following line converts Z4 to Z4:Z4
            nsobj.lchain = nsobj.lchain + ":" + nsobj.lchain
        chains["L"] = workbook.active[nsobj.lchain]
        if nsobj.lambdakeys is not None:
            begin = nsobj.lambdakeys.split(":")[0]
            end = nsobj.lambdakeys.split(":")[1]
            columndict["L"] = createExportDict(ws, begin, end)
            if "Comment" not in columndict["L"].keys():
                sys.exit(
                    "No column titled 'Comment' in the fields specified by --lambdakeys. Exiting..."
                )
        else:
            print("--lambdakeys not set. No lambda chain data will be written")

    if nsobj.kchain is not None:
        if (
            nsobj.kchain.find(":") == -1
        ):  # we suppose this a single cell then. The following line converts Z4 to Z4:Z4
            nsobj.kchain = nsobj.kchain + ":" + nsobj.kchain
        chains["K"] = workbook.active[nsobj.kchain]
        if nsobj.kappakeys is not None:
            begin = nsobj.kappakeys.split(":")[0]
            end = nsobj.kappakeys.split(":")[1]
            columndict["K"] = createExportDict(ws, begin, end)
            if "Comment" not in columndict["K"].keys():
                sys.exit(
                    "No column titled 'Comment' in the fields specified by --kappakeys. Exiting..."
                )
        else:
            print("--kappakeys not set. No lambda chain data will be written")

    parsed_sequences = []

    for ct in chains.keys():

        ###chains['H']] is loaded by chains['H']=workbook.active[nsobj.heavy]. if nsobj.heavy=Z (a whole row),
        ###a list is returned, but if nsobj.heavy=Z4:Z240 (for example..), then a tuple is returen (?!?).
        ### thats why we have to iterate over seq, instead of seq###
        for (seq,) in chains[ct]:
            if seq.value is None:
                continue

            # check if this line has already been analyzed - and skip, nsobj.overwrite is not set
            if seq.value is not None and nsobj.overwrite is False:
                if ws[columndict[ct]["Confirmation"] + str(seq.row)].value != None:
                    print(
                        ws[columndict[ct]["Confirmation"] + str(seq.row)].value
                        + " has already been analyzed."
                    )
                    continue

            filename = seq.value

            # 21.03.21 the following line is a workaround for the inconsistent naming scheme of Eurofins
            filename = filename.replace("-", "_")

            if nsobj.dataprefix is not None:
                filename = nsobj.dataprefix + str(filename) + ".ab1"
            try:
                # shutil.copy2(filename, "../results/", follow_symlinks=True)
                #
                my_ps = SequenceFile(filename)
                parsed_sequences.append(my_ps)

                if my_ps.successfullyParsed == False:
                    ws[columndict[ct]["Comment"] + str(seq.row)] = my_ps.comment
                    ws[columndict[ct]["QV"] + str(seq.row)] = my_ps.mean_phred_quality
                    ws[columndict[ct]["Confirmation"] + str(seq.row)] = (
                        "to be confirmed"
                    )
                    ws[columndict[ct]["Function"] + str(seq.row)] = "BQ"
                    ws[columndict[ct]["RL"] + str(seq.row)] = my_ps.len
                elif my_ps.chain_type is not ct:
                    ws[columndict[ct]["Comment"] + str(seq.row)] = (
                        my_ps.comment
                        + " "
                        + filename
                        + " has chain type "
                        + my_ps.chain_type
                    )
                    ws[columndict[ct]["QV"] + str(seq.row)] = my_ps.mean_phred_quality
                    ws[columndict[ct]["RL"] + str(seq.row)] = my_ps.len
                    ws[columndict[ct]["Confirmation"] + str(seq.row)] = (
                        "to be confirmed"
                    )
                    ws[columndict[ct]["Function"] + str(seq.row)] = "BQ"
                else:
                    updateExcelRow(workbook, seq.row, columndict[ct], my_ps)
            except FileNotFoundError:
                try:
                    ws[columndict[ct]["Function"] + str(seq.row)] = (
                        "BQ - file not found"
                    )
                    ws[columndict[ct]["Comment"] + str(seq.row)] = (
                        "File " + str(filename) + " not found."
                    )
                except:
                    sys.exit(
                        "OOPS! File "
                        + filename
                        + " not found! Also, no 'Comment' column was found. Please make sure there is such a column given in the range specified."
                    )
            except ValueError as my_err:
                sys.exit(
                    "OOPS! An error occured while parsing "
                    + filename
                    + ". "
                    + str(my_err)
                    + ". Aborting ..."
                )
            except OSError as my_err:
                sys.exit(
                    "OOPS! An error occured while parsing "
                    + filename
                    + ". "
                    + str(my_err)
                    + ". Maybe the wrong filetype? Aborting ..."
                )

    if nsobj.heavykeys is None or nsobj.lambdakeys is None or nsobj.kappakeys is None:
        sys.exit("Please set keys")

    workbook.save(nsobj.outfile)

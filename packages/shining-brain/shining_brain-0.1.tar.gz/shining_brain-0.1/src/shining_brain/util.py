import re
from pathlib import Path

import openpyxl
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.orm import declarative_base, sessionmaker
from sqlalchemy.sql import text

from shining_brain.config_loader import application_config
from shining_brain.database import CreateTableStatement, CreateDefinition, ColumnDefinition
from shining_brain.logger_setup import setup_logger

log = setup_logger("util.py")


def load_file_into_database(project, data_frame, table_name, column_mapping=None, before_statement=None, after_statement=None):
    data_frame = rename_columns(column_mapping, data_frame)

    save_data_frame(project, after_statement, before_statement, data_frame, table_name)


def rename_columns(column_mapping, data_frame):
    data_frame = data_frame[column_mapping.keys()]
    data_frame.rename(columns=column_mapping, inplace=True)
    return data_frame


def save_data_frame(project, after_statement, before_statement, data_frame, table_name):
    base = declarative_base()
    engine = create_engine(get_url(project))
    base.metadata.create_all(engine)
    session_maker = sessionmaker(bind=engine)
    session = session_maker()
    if before_statement is not None:
        session.execute(text(before_statement))
        session.commit()
    data_frame.to_sql(table_name, con=engine, index=False, if_exists='append')
    if after_statement is not None:
        session.execute(text(after_statement))
        session.commit()
        session.close()
    log.debug('%s successfully loaded into the database.', table_name)


def load_configurations(key):
    return application_config[key]


def get_url(project):
    database = project.get_database()
    return f'mysql+mysqlconnector://{database.get_user()}:{database.get_password()}@{database.get_host()}/{database.get_schema()}'


def to_snake_case(a_text):
    return re.sub(r'\W+', '_', a_text).lower()


def generate_create_table_statement(columns, mapping, table_name):
    create_definitions = []
    for column in columns:
        if column not in mapping[table_name]:
            continue
        create_definitions.append(CreateDefinition(to_snake_case(column), ColumnDefinition()))

    return CreateTableStatement(table_name, create_definitions).get()


def load_file_into_data_frame(file_path):
    if file_path.endswith('.csv'):
        return pd.read_csv(file_path)
    if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
        return pd.read_excel(file_path)
    raise ValueError("The application is capable of handling only CSV and Excel files.")


def generate_column_mapping(data_frame):
    column_mapping = {}
    for column in data_frame.columns:
        column_name = to_snake_case(column)
        column_mapping[column] = column_name

    return column_mapping


def gather_file_paths(directory_path, files, pattern):
    dir_path = Path(directory_path)
    if not dir_path.is_dir():
        raise ValueError(f"The path {directory_path} is not a valid directory")

    for file_path in dir_path.iterdir():
        if file_path.is_dir():
            gather_file_paths(file_path, files, pattern)
            continue
        if re.match(pattern, file_path.as_posix()):
            files.append(file_path)


def extract_colored_rows(xlsx, filter_colors):
    workbook = openpyxl.load_workbook(xlsx)
    df = pd.read_excel(xlsx, nrows=1)
    df['Color'] = None
    column_names = df.columns
    sheet = workbook.active
    colored_rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
        for cell in row:
            cell_color = cell.fill.start_color.index[2:]
            if cell_color in filter_colors:
                values = [c.value for c in row]
                values.append(cell_color)
                colored_rows.append(values)
                break
    result_df = pd.DataFrame(colored_rows)
    result_df.columns = column_names
    return result_df


def extract_non_black_rows_from_excel(io, header_name):
    wb = openpyxl.load_workbook(io)
    ws = wb.active

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if header_name in row:
            header_row = i
            break
    if header_row is None:
        return None
    df = pd.read_excel(io, header=header_row-1, nrows=1)
    column_names = df.columns
    filtered_rows = []
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, values_only=False):
        for cell in row:
            # Assuming the color is in ARGB format and use the standard ARGB code for black
            if cell.fill.start_color.index == '00000000':
                continue
            values = [c.value for c in row]
            filtered_rows.append(values)
            break
    if len(filtered_rows) == 0:
        return None
    filtered_df = pd.DataFrame(data=filtered_rows)
    filtered_df.columns = column_names
    return filtered_df
